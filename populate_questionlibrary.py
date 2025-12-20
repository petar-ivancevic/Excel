"""
Populate QuestionLibrary_shell.xlsx with synthetic test data for KYC onboarding prototype
"""
import openpyxl
from datetime import datetime, timedelta
import random

# Load the workbook
wb = openpyxl.load_workbook('QuestionLibrary_shell.xlsx')

print("Populating QuestionLibrary_shell.xlsx with synthetic data...\n")

# ============================================================================
# 1. JURISDICTIONS (tblJurisdictions) - 5 rows
# ============================================================================
ws_jurisdictions = wb['Jurisdictions']
jurisdictions_data = [
    ('US', 'United States'),
    ('UK', 'United Kingdom'),
    ('CA-ON', 'Ontario (Canada)'),
    ('SG', 'Singapore'),
    ('DE', 'Germany')
]

for idx, (jur_id, jur_name) in enumerate(jurisdictions_data, start=2):
    ws_jurisdictions.cell(row=idx, column=1, value=jur_id)
    ws_jurisdictions.cell(row=idx, column=2, value=jur_name)

print(f"[OK] Jurisdictions: {len(jurisdictions_data)} rows populated")

# ============================================================================
# 2. AUDITORS (tblAuditors) - 4 rows
# ============================================================================
ws_auditors = wb['Auditors']
auditors_data = [
    ('A001', 'John Smith', 'john.smith@example.com'),
    ('A002', 'Maria Lopez', 'maria.lopez@example.com'),
    ('A003', 'Anil Patel', 'anil.patel@example.com'),
    ('A004', 'Sophie Mueller', 'sophie.mueller@example.com')
]

for idx, (aud_id, name, email) in enumerate(auditors_data, start=2):
    ws_auditors.cell(row=idx, column=1, value=aud_id)
    ws_auditors.cell(row=idx, column=2, value=name)
    ws_auditors.cell(row=idx, column=3, value=email)

print(f"[OK] Auditors: {len(auditors_data)} rows populated")

# ============================================================================
# 3. CLIENT TYPE RISK (tblClientTypeRisk) - 7 rows
# ============================================================================
ws_clienttyperisk = wb['ClientTypeRisk']
clienttyperisk_data = [
    ('Hedge Fund', 'EDD', 'Y'),
    ('Money Service Business', 'EDD', 'Y'),
    ('Private Equity Fund', 'High', 'Y'),
    ('Financial Institution', 'High', 'N'),
    ('Listed Corporate', 'Standard', 'N'),
    ('Retail Individual', 'Standard', 'N'),
    ('Corporate', 'Standard', 'N')
]

for idx, (client_type, risk_tier, is_edd) in enumerate(clienttyperisk_data, start=2):
    ws_clienttyperisk.cell(row=idx, column=1, value=client_type)
    ws_clienttyperisk.cell(row=idx, column=2, value=risk_tier)
    ws_clienttyperisk.cell(row=idx, column=3, value=is_edd)

print(f"[OK] ClientTypeRisk: {len(clienttyperisk_data)} rows populated")

# ============================================================================
# 4. ENTITIES (tblEntities) - 30 rows
# ============================================================================
ws_entities = wb['Entities']

# Helper lists
jurisdiction_ids = ['US', 'UK', 'CA-ON', 'SG', 'DE']
jurisdiction_map = dict(jurisdictions_data)
client_types = [ct[0] for ct in clienttyperisk_data]
families = [
    (9001, 'Acme Group'),
    (9002, 'Maple Group'),
    (9003, 'Sunrise Group'),
    (9004, 'Pinnacle Group'),
    (9005, 'Horizon Group'),
    (9006, 'Vertex Group')
]
flu_list = ['CIB', 'CB', 'Markets', 'Retail']
irr_list = ['Low', 'Medium', 'High']
restriction_levels = ['L1', 'L2', 'L3']
oper_status_list = ['Active', 'Active', 'Active', 'Dormant', 'Closed']
party_types = ['Corporate', 'Financial Institution', 'Individual']
kyc_status_list = ['Complete', 'In Progress', 'Not Started']
drr_list = ['Low', 'Medium', 'High']

# Generate 30 entities
entities_data = []
gci_counter = 1001

for family_gci, family_name in families:
    # Each family has 3-6 entities
    num_entities = random.randint(3, 6)

    for i in range(num_entities):
        if gci_counter > 1030:
            break

        # Generate legal name variations
        suffixes = ['Holdings LLC', 'Treasury Ltd', 'Technologies Inc', 'Capital Partners', 'Investments LP', 'Group plc']
        legal_name = f"{family_name.replace(' Group', '')} {random.choice(suffixes)}"

        # Random KYC date between 2021-01-01 and 2024-12-31
        start_date = datetime(2021, 1, 1)
        end_date = datetime(2024, 12, 31)
        time_between = end_date - start_date
        random_days = random.randint(0, time_between.days)
        kyc_date = (start_date + timedelta(days=random_days)).strftime('%Y-%m-%d')

        # Select jurisdiction
        jur_id = random.choice(jurisdiction_ids)
        jur_name = jurisdiction_map[jur_id]

        # Country of incorporation (map from jurisdiction)
        country_map = {'US': 'US', 'UK': 'GB', 'CA-ON': 'CA', 'SG': 'SG', 'DE': 'DE'}
        country = country_map.get(jur_id, 'US')

        # Entity attributes
        entity = {
            'GCI': gci_counter,
            'Family GCI': family_gci,
            'Family Name': family_name,
            'Legal Name': legal_name,
            'KYC Date': kyc_date,
            'Jurisdiction': jur_name,
            'Jurisdiction ID': jur_id,
            'Bk. Entity': random.choice(['Y', 'Y', 'Y', 'N']),
            'Primary FLU': random.choice(flu_list),
            'IRR': random.choice(irr_list),
            'Juris. Status': random.choice(['Active', 'Active', 'Active', 'Inactive']),
            'Restriction Level BE Cd': random.choice(restriction_levels),
            'Oper. Status': random.choice(oper_status_list),
            'Party Type': random.choice(party_types),
            'KYC Status': random.choice(kyc_status_list),
            'DRR': random.choice(drr_list),
            'DRR Reason': random.choice(['', '', 'High value transactions', 'Complex structure', 'PEP association']),
            'Client Type': random.choice(client_types),
            'Refresh LOB': random.choice(['Markets', 'Retail', 'CIB', 'CB']),
            'Country of Incorp.': country,
            'Restriction Comment': random.choice(['', '', '', 'Enhanced monitoring required', 'Pending documentation'])
        }

        entities_data.append(entity)
        gci_counter += 1

        if gci_counter > 1030:
            break

# Write entities to sheet
for idx, entity in enumerate(entities_data, start=2):
    ws_entities.cell(row=idx, column=1, value=entity['GCI'])
    ws_entities.cell(row=idx, column=2, value=entity['Family GCI'])
    ws_entities.cell(row=idx, column=3, value=entity['Family Name'])
    ws_entities.cell(row=idx, column=4, value=entity['Legal Name'])
    ws_entities.cell(row=idx, column=5, value=entity['KYC Date'])
    ws_entities.cell(row=idx, column=6, value=entity['Jurisdiction'])
    ws_entities.cell(row=idx, column=7, value=entity['Jurisdiction ID'])
    ws_entities.cell(row=idx, column=8, value=entity['Bk. Entity'])
    ws_entities.cell(row=idx, column=9, value=entity['Primary FLU'])
    ws_entities.cell(row=idx, column=10, value=entity['IRR'])
    ws_entities.cell(row=idx, column=11, value=entity['Juris. Status'])
    ws_entities.cell(row=idx, column=12, value=entity['Restriction Level BE Cd'])
    ws_entities.cell(row=idx, column=13, value=entity['Oper. Status'])
    ws_entities.cell(row=idx, column=14, value=entity['Party Type'])
    ws_entities.cell(row=idx, column=15, value=entity['KYC Status'])
    ws_entities.cell(row=idx, column=16, value=entity['DRR'])
    ws_entities.cell(row=idx, column=17, value=entity['DRR Reason'])
    ws_entities.cell(row=idx, column=18, value=entity['Client Type'])
    ws_entities.cell(row=idx, column=19, value=entity['Refresh LOB'])
    ws_entities.cell(row=idx, column=20, value=entity['Country of Incorp.'])
    ws_entities.cell(row=idx, column=21, value=entity['Restriction Comment'])

print(f"[OK] Entities: {len(entities_data)} rows populated")

# ============================================================================
# 5. ATTRIBUTES (tblAttributes) - 30 rows
# ============================================================================
ws_attributes = wb['Attributes']

source_files = ['US_KYC.pdf', 'UK_KYC.pdf', 'CA_KYC.pdf', 'SG_KYC.pdf', 'DE_KYC.pdf']
categories = ['Entity Profile', 'Ownership', 'AML', 'EDD', 'Documentation']
sources = ['Policy 1', 'Policy 2', 'EDD Policy', 'AML Policy']

attributes_data = [
    # Base attributes
    ('US_KYC.pdf', 'A001', 'Legal Form & Registration', 'Entity Profile', 'Policy 1', 3,
     'Confirm the entity\'s legal form and registration details.', 'Core requirement', 'All', 'Base', 'Y', 365),
    ('UK_KYC.pdf', 'A002', 'Registered Address', 'Entity Profile', 'Policy 1', 4,
     'Verify the entity\'s registered address.', '', 'All', 'Base', 'Y', 365),
    ('CA_KYC.pdf', 'A003', 'Business Activities', 'Entity Profile', 'Policy 1', 5,
     'Describe the entity\'s primary business activities.', '', 'All', 'Base', 'Y', 730),
    ('US_KYC.pdf', 'A004', 'Beneficial Owners', 'Ownership', 'Policy 2', 8,
     'Identify all beneficial owners with 25% or greater ownership.', 'Critical', 'All', 'Base', 'Y', 365),
    ('SG_KYC.pdf', 'A005', 'Ultimate Beneficial Owner', 'Ownership', 'Policy 2', 9,
     'Identify the ultimate beneficial owner of the entity.', '', 'All', 'Base', 'Y', 365),
    ('UK_KYC.pdf', 'A006', 'Directors and Officers', 'Ownership', 'Policy 2', 10,
     'Provide details of all directors and senior officers.', '', 'All', 'Base', 'Y', 365),
    ('DE_KYC.pdf', 'A007', 'Organizational Chart', 'Ownership', 'Policy 2', 11,
     'Submit an organizational chart showing ownership structure.', '', 'All', 'Both', 'Y', 365),
    ('US_KYC.pdf', 'A008', 'Sanctions Screening', 'AML', 'AML Policy', 15,
     'Confirm sanctions screening has been completed for all parties.', '', 'All', 'Base', 'Y', None),
    ('CA_KYC.pdf', 'A009', 'PEP Status', 'AML', 'AML Policy', 16,
     'Determine if any beneficial owners are Politically Exposed Persons.', '', 'All', 'Base', 'Y', None),
    ('SG_KYC.pdf', 'A010', 'Adverse Media Check', 'AML', 'AML Policy', 17,
     'Complete adverse media checks on the entity and key individuals.', '', 'All', 'Base', 'Y', None),

    # EDD attributes
    ('US_KYC.pdf', 'A011', 'Source of Wealth (High Risk)', 'EDD', 'EDD Policy', 20,
     'Provide detailed documentation on the source of wealth for high-risk clients.', 'EDD only', 'All', 'EDD', 'Y', 365),
    ('UK_KYC.pdf', 'A012', 'Source of Funds', 'EDD', 'EDD Policy', 21,
     'Document the source of funds for all transactions.', 'EDD only', 'All', 'EDD', 'Y', 365),
    ('CA_KYC.pdf', 'A013', 'Enhanced Background Checks', 'EDD', 'EDD Policy', 22,
     'Conduct enhanced background checks on all beneficial owners.', 'EDD only', 'All', 'EDD', 'Y', 180),
    ('SG_KYC.pdf', 'A014', 'Business Rationale', 'EDD', 'EDD Policy', 23,
     'Provide detailed explanation of the business relationship rationale.', '', 'All', 'EDD', 'Y', 730),
    ('DE_KYC.pdf', 'A015', 'Enhanced Financial Analysis', 'EDD', 'EDD Policy', 24,
     'Submit audited financial statements and additional financial analysis.', 'EDD only', 'All', 'EDD', 'Y', 180),

    # Both Base and EDD
    ('US_KYC.pdf', 'A016', 'Financial Statements', 'Documentation', 'Policy 1', 12,
     'Provide most recent financial statements.', 'Annual requirement', 'All', 'Both', 'Y', 365),
    ('UK_KYC.pdf', 'A017', 'Tax Identification', 'Documentation', 'Policy 1', 13,
     'Submit tax identification number documentation.', '', 'All', 'Base', 'Y', 730),
    ('CA_KYC.pdf', 'A018', 'Bank References', 'Documentation', 'Policy 2', 14,
     'Obtain bank references from existing financial institutions.', '', 'All', 'Both', 'N', 365),
    ('SG_KYC.pdf', 'A019', 'Trade References', 'Documentation', 'Policy 2', 15,
     'Obtain professional or trade references.', '', 'All', 'Base', 'N', 730),
    ('DE_KYC.pdf', 'A020', 'Insurance Documentation', 'Documentation', 'Policy 1', 7,
     'Provide proof of professional indemnity insurance where applicable.', '', 'US', 'Base', 'N', 365),

    # Additional attributes
    ('US_KYC.pdf', 'A021', 'Regulatory Licenses', 'Entity Profile', 'Policy 1', 6,
     'Submit copies of all relevant regulatory licenses.', '', 'All', 'Base', 'Y', 365),
    ('UK_KYC.pdf', 'A022', 'Related Parties', 'Ownership', 'Policy 2', 12,
     'Disclose all related party relationships.', '', 'All', 'Both', 'Y', 365),
    ('CA_KYC.pdf', 'A023', 'Country Risk Assessment', 'AML', 'AML Policy', 18,
     'Assess country risk for all jurisdictions where the entity operates.', '', 'All', 'Both', 'Y', None),
    ('SG_KYC.pdf', 'A024', 'Transaction Monitoring', 'AML', 'AML Policy', 19,
     'Establish transaction monitoring parameters.', '', 'All', 'Base', 'Y', None),
    ('DE_KYC.pdf', 'A025', 'Client Interview', 'EDD', 'EDD Policy', 25,
     'Conduct in-person or video interview with senior management.', 'EDD only', 'All', 'EDD', 'Y', None),
    ('US_KYC.pdf', 'A026', 'Third Party Introductions', 'Documentation', 'Policy 2', 8,
     'Document any third party introducers and conduct due diligence.', '', 'All', 'Base', 'N', 365),
    ('UK_KYC.pdf', 'A027', 'Website Verification', 'Entity Profile', 'Policy 1', 5,
     'Verify the entity\'s website and online presence.', '', 'All', 'Base', 'N', None),
    ('CA_KYC.pdf', 'A028', 'Proof of Address', 'Documentation', 'Policy 1', 6,
     'Obtain proof of address for registered office.', '', 'All', 'Base', 'Y', 90),
    ('SG_KYC.pdf', 'A029', 'Relationship Purpose', 'Entity Profile', 'Policy 1', 7,
     'Document the anticipated purpose and nature of the business relationship.', '', 'All', 'Base', 'Y', 730),
    ('DE_KYC.pdf', 'A030', 'Periodic Review Schedule', 'Documentation', 'Policy 1', 9,
     'Establish periodic review schedule based on risk rating.', '', 'All', 'Both', 'Y', None)
]

for idx, attr in enumerate(attributes_data, start=2):
    ws_attributes.cell(row=idx, column=1, value=attr[0])   # Source File
    ws_attributes.cell(row=idx, column=2, value=attr[1])   # Attribute ID
    ws_attributes.cell(row=idx, column=3, value=attr[2])   # Attribute Name
    ws_attributes.cell(row=idx, column=4, value=attr[3])   # Category
    ws_attributes.cell(row=idx, column=5, value=attr[4])   # Source
    ws_attributes.cell(row=idx, column=6, value=attr[5])   # Source Page
    ws_attributes.cell(row=idx, column=7, value=attr[6])   # Question Text
    ws_attributes.cell(row=idx, column=8, value=attr[7])   # Notes
    ws_attributes.cell(row=idx, column=9, value=attr[8])   # Jurisdiction ID
    ws_attributes.cell(row=idx, column=10, value=attr[9])  # RiskScope
    ws_attributes.cell(row=idx, column=11, value=attr[10]) # IsRequired
    ws_attributes.cell(row=idx, column=12, value=attr[11]) # DocumentationAgeRule

print(f"[OK] Attributes: {len(attributes_data)} rows populated")

# ============================================================================
# 6. ACCEPTABLE DOCS (tblAcceptableDocs) - ~60 rows
# ============================================================================
ws_acceptabledocs = wb['AcceptableDocs']

# Document evidence for each attribute (1-3 docs per attribute)
acceptable_docs_data = [
    # A001 - Legal Form & Registration
    ('US_KYC.pdf', 'A001', 'Registration', 'Certificate of Incorporation', 'All', 'Must be certified'),
    ('US_KYC.pdf', 'A001', 'Registration', 'Articles of Association', 'All', ''),

    # A002 - Registered Address
    ('UK_KYC.pdf', 'A002', 'Registration', 'Certificate of Good Standing', 'All', ''),
    ('UK_KYC.pdf', 'A002', 'Other', 'Utility Bill (Registered Office)', 'All', 'Within 3 months'),

    # A003 - Business Activities
    ('CA_KYC.pdf', 'A003', 'Other', 'Business License', 'All', ''),
    ('CA_KYC.pdf', 'A003', 'Other', 'Company Website Print', 'All', ''),

    # A004 - Beneficial Owners
    ('US_KYC.pdf', 'A004', 'Ownership', 'Share Register', 'All', 'Certified'),
    ('US_KYC.pdf', 'A004', 'Ownership', 'Shareholder Declaration', 'All', ''),
    ('US_KYC.pdf', 'A004', 'Identification', 'Passport Copy (Beneficial Owner)', 'All', 'Government issued'),

    # A005 - Ultimate Beneficial Owner
    ('SG_KYC.pdf', 'A005', 'Ownership', 'UBO Declaration Form', 'All', 'Signed and dated'),
    ('SG_KYC.pdf', 'A005', 'Identification', 'National ID Card (UBO)', 'All', ''),

    # A006 - Directors and Officers
    ('UK_KYC.pdf', 'A006', 'Registration', 'Register of Directors', 'All', 'From registry'),
    ('UK_KYC.pdf', 'A006', 'Identification', 'Passport Copy (Directors)', 'All', ''),

    # A007 - Organizational Chart
    ('DE_KYC.pdf', 'A007', 'Ownership', 'Ownership Structure Diagram', 'All', 'Must show percentages'),
    ('DE_KYC.pdf', 'A007', 'Ownership', 'Group Structure Chart', 'All', ''),

    # A008 - Sanctions Screening
    ('US_KYC.pdf', 'A008', 'Other', 'Sanctions Screening Report', 'All', 'Completed screening'),

    # A009 - PEP Status
    ('CA_KYC.pdf', 'A009', 'Other', 'PEP Screening Report', 'All', ''),
    ('CA_KYC.pdf', 'A009', 'Other', 'PEP Declaration Form', 'All', ''),

    # A010 - Adverse Media Check
    ('SG_KYC.pdf', 'A010', 'Other', 'Adverse Media Report', 'All', 'Independent source'),

    # A011 - Source of Wealth (EDD)
    ('US_KYC.pdf', 'A011', 'EDD', 'Source of Wealth Statement', 'All', 'EDD requirement'),
    ('US_KYC.pdf', 'A011', 'EDD', 'Asset Documentation', 'All', 'Supporting evidence'),
    ('US_KYC.pdf', 'A011', 'EDD', 'Tax Returns (3 years)', 'All', ''),

    # A012 - Source of Funds (EDD)
    ('UK_KYC.pdf', 'A012', 'EDD', 'Source of Funds Declaration', 'All', 'EDD requirement'),
    ('UK_KYC.pdf', 'A012', 'EDD', 'Bank Statements (6 months)', 'All', ''),

    # A013 - Enhanced Background Checks (EDD)
    ('CA_KYC.pdf', 'A013', 'EDD', 'Enhanced Background Check Report', 'All', 'Third party verification'),
    ('CA_KYC.pdf', 'A013', 'EDD', 'Criminal Record Check', 'All', 'Where applicable'),

    # A014 - Business Rationale (EDD)
    ('SG_KYC.pdf', 'A014', 'EDD', 'Business Rationale Memo', 'All', 'Detailed explanation'),

    # A015 - Enhanced Financial Analysis (EDD)
    ('DE_KYC.pdf', 'A015', 'EDD', 'Audited Financial Statements (3 years)', 'All', 'Must be audited'),
    ('DE_KYC.pdf', 'A015', 'EDD', 'Financial Analysis Report', 'All', ''),

    # A016 - Financial Statements
    ('US_KYC.pdf', 'A016', 'Other', 'Annual Financial Statements', 'All', 'Most recent year'),
    ('US_KYC.pdf', 'A016', 'Other', 'Management Accounts', 'All', 'If recent year-end not available'),

    # A017 - Tax Identification
    ('UK_KYC.pdf', 'A017', 'Registration', 'Tax Registration Certificate', 'All', ''),
    ('UK_KYC.pdf', 'A017', 'Other', 'W-9 Form', 'US', 'US entities only'),

    # A018 - Bank References
    ('CA_KYC.pdf', 'A018', 'Other', 'Bank Reference Letter', 'All', 'On bank letterhead'),

    # A019 - Trade References
    ('SG_KYC.pdf', 'A019', 'Other', 'Professional Reference Letter', 'All', ''),
    ('SG_KYC.pdf', 'A019', 'Other', 'Trade Reference Letter', 'All', ''),

    # A020 - Insurance Documentation
    ('DE_KYC.pdf', 'A020', 'Other', 'Professional Indemnity Insurance Certificate', 'US', ''),

    # A021 - Regulatory Licenses
    ('US_KYC.pdf', 'A021', 'Registration', 'Regulatory License Copy', 'All', 'Certified copy'),
    ('US_KYC.pdf', 'A021', 'Registration', 'License Renewal Confirmation', 'All', ''),

    # A022 - Related Parties
    ('UK_KYC.pdf', 'A022', 'Ownership', 'Related Parties Disclosure', 'All', ''),

    # A023 - Country Risk Assessment
    ('CA_KYC.pdf', 'A023', 'Other', 'Country Risk Analysis Report', 'All', ''),

    # A024 - Transaction Monitoring
    ('SG_KYC.pdf', 'A024', 'Other', 'Transaction Monitoring Parameters', 'All', ''),

    # A025 - Client Interview (EDD)
    ('DE_KYC.pdf', 'A025', 'EDD', 'Interview Minutes', 'All', 'EDD requirement'),
    ('DE_KYC.pdf', 'A025', 'EDD', 'Interview Checklist', 'All', ''),

    # A026 - Third Party Introductions
    ('US_KYC.pdf', 'A026', 'Other', 'Introducer Due Diligence', 'All', ''),

    # A027 - Website Verification
    ('UK_KYC.pdf', 'A027', 'Other', 'Website Screenshots', 'All', ''),

    # A028 - Proof of Address
    ('CA_KYC.pdf', 'A028', 'Other', 'Utility Bill', 'All', 'Within 90 days'),
    ('CA_KYC.pdf', 'A028', 'Other', 'Lease Agreement', 'All', ''),

    # A029 - Relationship Purpose
    ('SG_KYC.pdf', 'A029', 'Other', 'Relationship Purpose Questionnaire', 'All', 'Completed form'),

    # A030 - Periodic Review Schedule
    ('DE_KYC.pdf', 'A030', 'Other', 'Review Schedule Documentation', 'All', 'Risk-based'),

    # Additional docs to reach ~60 rows
    ('US_KYC.pdf', 'A001', 'Registration', 'Memorandum of Association', 'UK', 'UK entities'),
    ('UK_KYC.pdf', 'A004', 'Ownership', 'Trust Deed', 'All', 'If applicable'),
    ('CA_KYC.pdf', 'A007', 'Ownership', 'Partnership Agreement', 'All', 'For partnerships'),
    ('SG_KYC.pdf', 'A016', 'Other', 'Interim Financial Statements', 'All', 'If recent changes'),
    ('DE_KYC.pdf', 'A011', 'EDD', 'Wealth Verification Letter', 'All', 'From accountant'),
    ('US_KYC.pdf', 'A021', 'Registration', 'Registration Confirmation Letter', 'All', 'From regulator'),
    ('UK_KYC.pdf', 'A006', 'Identification', 'Proof of Address (Directors)', 'All', ''),
    ('CA_KYC.pdf', 'A005', 'Ownership', 'UBO Certification', 'All', 'Notarized'),
]

for idx, doc in enumerate(acceptable_docs_data, start=2):
    ws_acceptabledocs.cell(row=idx, column=1, value=doc[0])  # Source File
    ws_acceptabledocs.cell(row=idx, column=2, value=doc[1])  # Attribute ID
    ws_acceptabledocs.cell(row=idx, column=3, value=doc[2])  # Evidence Type
    ws_acceptabledocs.cell(row=idx, column=4, value=doc[3])  # Evidence Source/Document
    ws_acceptabledocs.cell(row=idx, column=5, value=doc[4])  # Jurisdiction ID
    ws_acceptabledocs.cell(row=idx, column=6, value=doc[5])  # Notes

print(f"[OK] AcceptableDocs: {len(acceptable_docs_data)} rows populated")

# ============================================================================
# Save the workbook
# ============================================================================
wb.save('QuestionLibrary_shell.xlsx')
print("\n[SUCCESS] Successfully saved QuestionLibrary_shell.xlsx")
print("\n" + "="*60)
print("SUMMARY - Data rows populated per table:")
print("="*60)
print(f"Jurisdictions:     {len(jurisdictions_data)} rows")
print(f"Auditors:          {len(auditors_data)} rows")
print(f"ClientTypeRisk:    {len(clienttyperisk_data)} rows")
print(f"Entities:          {len(entities_data)} rows")
print(f"Attributes:        {len(attributes_data)} rows")
print(f"AcceptableDocs:    {len(acceptable_docs_data)} rows")
print("="*60)
